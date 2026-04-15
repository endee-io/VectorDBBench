#!/usr/bin/env python3
"""
Normal Case Benchmark Script - Qdrant - Performance768D10M
(vectors pre-loaded, no load step)

Workflow:
  Run vectordbbench qdrantlocal for top-k values:
    30, 62, 64, 100, 128, 256, 512, 1000
  Each run is executed FOUR times; the result with the highest QPS is recorded.
  10-second gap between every run.

Output: Excel file with results.
"""

import subprocess
import json
import os
import time
import glob as glob_module
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# ============================================================
# CONFIGURATION
# ============================================================
QDRANT_URL        = "http://148.113.54.173:6333"
DATASET_LOCAL_DIR = "/home/debian/latest_VDB/VectorDBBench/vectordataset_label_10M"
RESULTS_DIR       = "/home/debian/latest_VDB/VectorDBBench/vectordb_bench/results/QdrantLocal"
TASK_LABEL_PREFIX = "qdrant_normal"
NUM_PER_BATCH     = 10000

M               = 16
HNSW_EF         = 128
EF_CONSTRUCT    = 128
CONCURRENCY     = 16
CONCURRENCY_DUR = 30
PRECISION       = "float32"
SPACE_TYPE      = "Cosine"
DATABASE        = "Qdrant"

TOP_K_VALUES = [30, 62, 64, 100, 128, 256, 512, 1000]

OUTPUT_EXCEL = os.path.join(
    "/home/debian/latest_VDB/VectorDBBench",
    f"normal_bench_qdrant_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
)

# ============================================================
# BENCHMARK RUNNER
# ============================================================

def run_vectordbbench(top_k: int) -> dict:
    before = set(glob_module.glob(os.path.join(RESULTS_DIR, "*.json")))
    task_label = f"{TASK_LABEL_PREFIX}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    cmd = (
        f'NUM_PER_BATCH={NUM_PER_BATCH} '
        f'DATASET_LOCAL_DIR="{DATASET_LOCAL_DIR}" '
        f'vectordbbench qdrantlocal '
        f'--url {QDRANT_URL} '
        f'--task-label {task_label} '
        f'--m {M} '
        f'--ef-construct {EF_CONSTRUCT} '
        f'--hnsw-ef {HNSW_EF} '
        f'--case-type Performance768D10M '
        f'--k {top_k} '
        f'--num-concurrency {CONCURRENCY} '
        f'--concurrency-duration {CONCURRENCY_DUR} '
        f'--concurrency-timeout 3600 '
        f'--skip-drop-old '
        f'--skip-load '
        f'--search-concurrent '
        f'--search-serial'
    )

    label = f"top_k={top_k}"
    print(f"\n  [RUN] {label}")
    proc = subprocess.run(cmd, shell=True, text=True)
    if proc.returncode != 0:
        print(f"  [WARN] vectordbbench exited with code {proc.returncode}")

    time.sleep(5)
    after = set(glob_module.glob(os.path.join(RESULTS_DIR, "*.json")))
    new_files = after - before

    if not new_files:
        print(f"  [ERROR] No new result file for {label}")
        return {"recall": None, "qps": None, "p99_latency": None, "load_duration": None}

    result_file = max(new_files, key=os.path.getmtime)
    print(f"  [FILE] {os.path.basename(result_file)}")

    with open(result_file) as f:
        data = json.load(f)

    metrics  = data["results"][0]["metrics"]
    recall   = metrics.get("recall")
    qps      = metrics.get("qps")
    p99      = metrics.get("serial_latency_p99")
    load_dur = metrics.get("load_duration")

    print(f"  [METRICS] recall={recall}, qps={qps}, p99={p99}, load_duration={load_dur}")
    return {"recall": recall, "qps": qps, "p99_latency": p99, "load_duration": load_dur}


def run_best_of_four(top_k: int) -> dict:
    """Run the benchmark four times and return the result with the highest QPS."""
    results = []
    for attempt in range(1, 5):
        print(f"\n  [ATTEMPT {attempt}/4] top_k={top_k}")
        result = run_vectordbbench(top_k)
        results.append(result)
        if attempt < 4:
            print(f"\n  [WAIT] 10s before next attempt ...")
            time.sleep(10)

    best = max(results, key=lambda r: r.get("qps") or 0)
    best_attempt = results.index(best) + 1
    print(f"  [BEST] Attempt {best_attempt} wins: qps={best.get('qps')}")
    return best


# ============================================================
# EXCEL WRITER
# ============================================================

def write_excel(rows: list, output_path: str):
    """
    rows: list of row_dict with: top_k, recall, qps, p99_latency, load_duration
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Normal Bench Qdrant"

    thin   = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")

    HDR_BG   = "1B4F72"
    HDR_FONT = Font(bold=True, color="FFFFFF")
    ROW_ODD  = "EBF5FB"
    ROW_EVEN = "FFFFFF"

    DATASET_NAME = "10M Cohere"

    columns = [
        ("Dataset",              20),
        ("Precision",            12),
        ("ef_con",               10),
        ("ef_search",            11),
        ("Concurrency",          14),
        ("top-k",                 8),
        ("Space Type",           12),
        ("Database",             12),
        ("Recall",               10),
        ("QPS",                  12),
        ("Latency (p99) in sec", 22),
        ("Load duration",        18),
    ]

    NUM_COLS = len(columns)

    for col_idx, (_, width) in enumerate(columns, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    current_row = 1

    # Title row
    title_cell = ws.cell(row=current_row, column=1,
                         value="Qdrant Normal Benchmark — Performance768D10M (best of 4 runs)")
    title_cell.font      = Font(bold=True, size=12)
    title_cell.alignment = left
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row,   end_column=NUM_COLS)
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    # Header row
    ws.row_dimensions[current_row].height = 28
    for col_idx, (header, _) in enumerate(columns, start=1):
        c = ws.cell(row=current_row, column=col_idx, value=header)
        c.font      = HDR_FONT
        c.fill      = PatternFill("solid", fgColor=HDR_BG)
        c.alignment = center
        c.border    = border
    current_row += 1

    # Data rows
    for row_local_idx, r in enumerate(rows):
        ws.row_dimensions[current_row].height = 22
        bg = ROW_ODD if row_local_idx % 2 == 0 else ROW_EVEN
        rf = PatternFill("solid", fgColor=bg)

        recall   = r.get("recall")
        qps      = r.get("qps")
        p99      = r.get("p99_latency")
        load_dur = r.get("load_duration")

        values = [
            DATASET_NAME,
            PRECISION,
            EF_CONSTRUCT,
            HNSW_EF,
            CONCURRENCY,
            r["top_k"],
            SPACE_TYPE,
            DATABASE,
            round(recall * 100, 2)  if recall   is not None else "N/A",
            round(qps, 4)           if qps       is not None else "N/A",
            round(p99, 6)           if p99       is not None else "N/A",
            round(load_dur, 4)      if load_dur  is not None else "N/A",
        ]

        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=current_row, column=col_idx, value=val)
            c.fill      = rf
            c.alignment = center if col_idx != 1 else left
            c.border    = border

        current_row += 1

    wb.save(output_path)
    print(f"\n[EXCEL] Saved → {output_path}")


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 60)
    print("Qdrant Normal Benchmark (best of 4 runs per top-k)")
    print(f"Qdrant URL : {QDRANT_URL}")
    print(f"Top-K      : {TOP_K_VALUES}")
    print(f"Output     : {OUTPUT_EXCEL}")
    print("=" * 60)

    rows = []

    for top_k in TOP_K_VALUES:
        metrics = run_best_of_four(top_k)
        rows.append({
            "top_k": top_k,
            **metrics,
        })

        print(f"\n  [WAIT] 10s before next top-k ...")
        time.sleep(10)

    write_excel(rows, OUTPUT_EXCEL)


if __name__ == "__main__":
    main()
