#!/usr/bin/env python3
"""
Stability Test Script - LTE Filter Operator
Index: (set INDEX_NAME below — 1M vectors pre-loaded, no load step)

Workflow:
  1. Fresh run     -> 4 filter rates (0.99, 0.80, 0.50, 0.01)
  2. Delete 1p     -> verify count (990000) -> bench 4 rates
  3. Wait 10s      -> Reinsert 1p  -> verify count (1000000) -> bench 4 rates
  4. Wait 10s      -> Delete 50p   -> verify count (500000)  -> bench 4 rates
  5. Wait 10s      -> Reinsert 50p -> verify count (1000000) -> bench 4 rates
  6. Wait 10s      -> Delete 80p   -> verify count (200000)  -> bench 4 rates
  7. Wait 10s      -> Reinsert 80p -> verify count (1000000) -> bench 4 rates
  8. Wait 10s      -> Delete 99p   -> verify count (10000)   -> bench 4 rates
  9. Wait 10s      -> Reinsert 99p -> verify count (1000000) -> bench 4 rates

Recorded per bench run: Recall (%), QPS, P99 Latency (s)
Recorded per operation: Delete duration (s), Insert duration (s)
Output: Excel file with all results
"""

import subprocess
import json
import os
import time
import glob as glob_module
import pyarrow.parquet as pq
from endee import Endee as EndeeClient
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# ============================================================
# CONFIGURATION
# ============================================================
TOKEN = "localtest"
BASE_URL = "http://148.113.54.173:8080/api/v1"
INDEX_NAME = "test_1M_vaib_filter_1503_adup3"
DATASET_LOCAL_DIR = "/home/debian/latest_VDB/VectorDBBench/vectordataset_lte"
BASE_PATH = "/home/debian/latest_VDB/VectorDBBench/vectordataset_lte"
DATASET_FOLDER = "cohere/cohere_medium_1m"
RESULTS_DIR = "/home/debian/latest_VDB/VectorDBBench/vectordb_bench/results/Endee"

FILTER_RATES = [0.99, 0.80, 0.50, 0.01]
DELETE_RATES = ["1p", "50p", "80p", "99p"]

# For delete: delete vectors with id >= lte_threshold (removes top portion)
RANGE_MAP_DELETE = {
    "1p":  990000,   # delete id >= 990000 → removes top 1%  → keeps 990,000
    "50p": 500000,   # delete id >= 500000 → removes top 50% → keeps 500,000
    "80p": 200000,   # delete id >= 200000 → removes top 80% → keeps 200,000
    "99p": 10000,    # delete id >= 10000  → removes top 99% → keeps 10,000
}

# For reinsert: reinsert vectors with id in [low, high]
RANGE_MAP_INSERT = {
    "1p":  [990000, 999999],   # reinsert top 1%
    "50p": [500000, 999999],   # reinsert top 50%
    "80p": [200000, 999999],   # reinsert top 80%
    "99p": [10000, 999999],    # reinsert top 99%
}

# Expected total_elements count after delete
EXPECTED_COUNT_AFTER_DELETE = {
    "1p":  990000,
    "50p": 500000,
    "80p": 200000,
    "99p": 10000,
}

TOTAL_VECTORS = 1000000

BATCH_SIZE = 1000

OUTPUT_EXCEL = os.path.join(
    "/home/debian/latest_VDB/VectorDBBench",
    f"stability_test_lte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
)

# ============================================================
# VECTORDBBENCH RUNNER
# ============================================================

def run_vectordbbench(filter_rate: float, prefilter: int | None = None) -> dict:
    """Run vectordbbench for one filter rate; return parsed metrics dict."""
    before = set(glob_module.glob(os.path.join(RESULTS_DIR, "*.json")))

    prefilter_flag = f'--prefilter-cardinality-threshold {prefilter} ' if prefilter else ''

    cmd = (
        f'DATASET_LOCAL_DIR="{DATASET_LOCAL_DIR}" vectordbbench endee '
        f'--token "{TOKEN}" '
        f'--region india-west-1 '
        f'--base-url "{BASE_URL}" '
        f'--index-name {INDEX_NAME} '
        f'--task-label "20260107" '
        f'--m 16 '
        f'--ef-con 128 '
        f'--ef-search 128 '
        f'--space-type cosine '
        f'{prefilter_flag}'
        f'--precision int16 '
        f'--version 1 '
        f'--case-type NewIntFilterPerformanceCase '
        f'--dataset-with-size-type "Medium Cohere (768dim, 1M)" '
        f'--filter-rate {filter_rate} '
        f'--k 30 '
        f'--num-concurrency "16" '
        f'--concurrency-duration 30 '
        f'--concurrency-timeout 3600 '
        f'--skip-drop-old '
        f'--skip-load '
        f'--search-concurrent '
        f'--search-serial'
    )

    prefilter_info = f", prefilter={prefilter}" if prefilter else ""
    print(f"\n  [RUN] filter_rate={filter_rate}{prefilter_info}")
    proc = subprocess.run(cmd, shell=True, text=True)
    if proc.returncode != 0:
        print(f"  [WARN] vectordbbench exited with code {proc.returncode}")

    time.sleep(3)
    after = set(glob_module.glob(os.path.join(RESULTS_DIR, "*.json")))
    new_files = after - before

    if not new_files:
        print(f"  [ERROR] No new result file found for filter_rate={filter_rate}")
        return {"recall": None, "qps": None, "p99_latency": None}

    result_file = max(new_files, key=os.path.getmtime)
    print(f"  [FILE] {os.path.basename(result_file)}")

    with open(result_file) as f:
        data = json.load(f)

    metrics = data["results"][0]["metrics"]
    recall = metrics.get("recall")
    qps    = metrics.get("qps")
    p99    = metrics.get("serial_latency_p99")

    print(f"  [METRICS] recall={recall}, qps={qps}, p99={p99}")
    return {"recall": recall, "qps": qps, "p99_latency": p99}


def run_all_filter_rates(label: str, prefilter: int | None = None) -> dict:
    """Run vectordbbench for all 4 filter rates. Returns {filter_rate: metrics}."""
    print(f"\n==> Running all filter rates [{label}]")
    results = {}
    for fr in FILTER_RATES:
        results[fr] = run_vectordbbench(fr, prefilter=prefilter)
        print("  [WAIT] 5s gap before next filter rate run ...")
        time.sleep(5)
    return results


# ============================================================
# DELETE / VERIFY / REINSERT
# ============================================================

def delete_vectors(index, rate_key: str) -> float:
    """Delete vectors with id >= threshold (removes top portion). Returns elapsed seconds."""
    threshold = RANGE_MAP_DELETE[rate_key]
    print(f"\n==> DELETE {rate_key}: removing vectors with id >= {threshold}")
    start = time.perf_counter()
    result = index.delete_with_filter([{"id": {"$gte": threshold}}])
    elapsed = time.perf_counter() - start
    print(f"  [DELETE] Completed in {elapsed:.2f}s | result: {result}")
    print("  [WAIT] 5s pause before verify count ...")
    time.sleep(5)
    return elapsed


def verify_count(index, expected: int) -> int:
    """Call index.describe() and print count vs expected."""
    desc = index.describe()
    actual = desc.get("count", "UNKNOWN")
    status = "OK" if actual == expected else "MISMATCH"
    print(f"  [VERIFY] Expected={expected}  Actual={actual}  [{status}]")
    return actual


def reinsert_vectors(index, rate_key: str) -> float:
    """Reinsert vectors in the deleted id range. Returns elapsed seconds."""
    low, high = RANGE_MAP_INSERT[rate_key]
    emb_path = os.path.join(BASE_PATH, DATASET_FOLDER, "shuffle_train.parquet")
    print(f"\n==> REINSERT {rate_key}: upserting vectors with id in [{low}, {high}]")

    emb_file = pq.ParquetFile(emb_path)
    total = 0
    start = time.perf_counter()

    for batch in emb_file.iter_batches(batch_size=BATCH_SIZE):
        df = batch.to_pandas()
        df = df[(df["id"] >= low) & (df["id"] <= high)]
        if df.empty:
            continue

        vectors = [
            {
                "id": str(row["id"]),
                "vector": row["emb"].tolist() if hasattr(row["emb"], "tolist") else list(row["emb"]),
                "meta": {"id": row["id"]},
                "filter": {"id": row["id"]},
            }
            for _, row in df.iterrows()
        ]
        index.upsert(vectors)
        total += len(vectors)
        print(f"  [INSERT] upserted {len(vectors)}, cumulative: {total}")

    elapsed = time.perf_counter() - start
    print(f"  [INSERT] Done. Total={total} vectors in {elapsed:.2f}s")
    return elapsed


# ============================================================
# EXCEL WRITER
# ============================================================

def write_excel(all_data: dict, output_path: str):
    """Write all collected results to a formatted Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stability Test LTE"

    def hdr(bold=True, color="FFFFFF"):
        return Font(bold=bold, color=color)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")
    thin   = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    COLOR_DARK    = "2E4057"
    COLOR_BLUE    = "1565C0"
    COLOR_LBLUE   = "BBDEFB"
    COLOR_GREEN   = "1B5E20"
    COLOR_LGREEN  = "C8E6C9"
    COLOR_ROW_ODD = "F5F5F5"

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 28

    # Phase cell (spans rows 1-2)
    ws.merge_cells("A1:A2")
    c = ws["A1"]
    c.value = "Phase"
    c.font = hdr(); c.fill = fill(COLOR_DARK)
    c.alignment = center; c.border = border

    # Filter rate group headers
    fr_labels = {0.99: "Filter Rate 0.99", 0.80: "Filter Rate 0.80",
                 0.50: "Filter Rate 0.50", 0.01: "Filter Rate 0.01"}
    col = 2
    for fr in FILTER_RATES:
        ws.merge_cells(
            start_row=1, start_column=col,
            end_row=1, end_column=col + 2
        )
        c = ws.cell(row=1, column=col, value=fr_labels[fr])
        c.font = hdr(); c.fill = fill(COLOR_BLUE)
        c.alignment = center; c.border = border
        col += 3

    # Operation duration header
    ws.merge_cells(
        start_row=1, start_column=col,
        end_row=1, end_column=col + 1
    )
    c = ws.cell(row=1, column=col, value="Operation Duration")
    c.font = hdr(); c.fill = fill(COLOR_GREEN)
    c.alignment = center; c.border = border

    # Row 2: Sub-headers
    col = 2
    for _ in FILTER_RATES:
        for sub in ("Recall (%)", "QPS", "P99 Latency (s)"):
            c = ws.cell(row=2, column=col, value=sub)
            c.font = Font(bold=True, color="000000")
            c.fill = fill(COLOR_LBLUE)
            c.alignment = center; c.border = border
            col += 1

    for sub in ("Delete Duration (s)", "Insert Duration (s)"):
        c = ws.cell(row=2, column=col, value=sub)
        c.font = Font(bold=True, color="000000")
        c.fill = fill(COLOR_LGREEN)
        c.alignment = center; c.border = border
        col += 1

    # Data rows
    phases = [
        ("Fresh Run\n(1M vectors, no delete)", "fresh", None, None),
    ]
    for rk in DELETE_RATES:
        exp_del = EXPECTED_COUNT_AFTER_DELETE[rk]
        phases.append((
            f"After Delete {rk}\n(verify count={exp_del:,})",
            f"post_delete_{rk}", f"delete_{rk}", None
        ))
        if rk == "99p":
            phases.append((
                "After Delete 99p\n(prefilter_cardinality_threshold=20000)",
                "post_delete_99p_prefilter", None, None
            ))
        phases.append((
            f"After Reinsert {rk}\n(verify count={TOTAL_VECTORS:,})",
            f"post_insert_{rk}", None, f"insert_{rk}"
        ))

    for i, (phase_name, bench_key, del_key, ins_key) in enumerate(phases):
        row = 3 + i
        ws.row_dimensions[row].height = 32
        row_fill = fill(COLOR_ROW_ODD) if i % 2 == 0 else fill("FFFFFF")

        c = ws.cell(row=row, column=1, value=phase_name)
        c.font = Font(bold=True); c.fill = row_fill
        c.alignment = left; c.border = border

        bench_results = all_data.get(bench_key) or {}
        col = 2
        for fr in FILTER_RATES:
            m = bench_results.get(fr) or {}
            recall = m.get("recall")
            qps    = m.get("qps")
            p99    = m.get("p99_latency")

            vals = [
                round(recall * 100, 2) if recall is not None else "N/A",
                round(qps, 4)          if qps    is not None else "N/A",
                round(p99, 6)          if p99    is not None else "N/A",
            ]
            for v in vals:
                c = ws.cell(row=row, column=col, value=v)
                c.fill = row_fill; c.alignment = center; c.border = border
                col += 1

        del_dur = all_data.get(del_key) if del_key else None
        ins_dur = all_data.get(ins_key) if ins_key else None

        c = ws.cell(row=row, column=col,
                    value=round(del_dur, 2) if del_dur is not None else "")
        c.fill = row_fill; c.alignment = center; c.border = border

        c = ws.cell(row=row, column=col + 1,
                    value=round(ins_dur, 2) if ins_dur is not None else "")
        c.fill = row_fill; c.alignment = center; c.border = border

    # Column widths
    ws.column_dimensions["A"].width = 30
    metric_cols = ["B","C","D","E","F","G","H","I","J","K","L","M"]
    for col_letter in metric_cols:
        ws.column_dimensions[col_letter].width = 15
    ws.column_dimensions["N"].width = 22
    ws.column_dimensions["O"].width = 22

    ws.freeze_panes = "B3"

    wb.save(output_path)
    print(f"\n[EXCEL] Saved: {output_path}")


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 65)
    print(" Stability Test — LTE Filter Operator")
    print(f" Index : {INDEX_NAME}")
    print(f" Output: {OUTPUT_EXCEL}")
    print("=" * 65)

    client = EndeeClient(token=TOKEN)
    client.set_base_url(BASE_URL)
    index = client.get_index(INDEX_NAME)
    print(f"\n[INIT] Connected to index '{INDEX_NAME}'")
    print(f"[INIT] Current state: {index.describe()}")

    all_data: dict = {}

    # ---- Phase 1: Fresh run ----
    all_data["fresh"] = run_all_filter_rates("Fresh — 1M vectors")

    for rate_key in DELETE_RATES:
        # ---- Delete ----
        all_data[f"delete_{rate_key}"] = delete_vectors(index, rate_key)

        # Verify count after delete
        verify_count(index, expected=EXPECTED_COUNT_AFTER_DELETE[rate_key])

        # ---- Bench post-delete ----
        all_data[f"post_delete_{rate_key}"] = run_all_filter_rates(
            f"Post-Delete {rate_key}"
        )

        # ---- Extra bench with prefilter=20000 only after 99p delete ----
        if rate_key == "99p":
            all_data["post_delete_99p_prefilter"] = run_all_filter_rates(
                "Post-Delete 99p (prefilter_cardinality_threshold=20000)",
                prefilter=20000
            )

        # Wait before reinsert
        print(f"\n[WAIT] 10s pause before reinsert {rate_key} ...")
        time.sleep(10)

        # ---- Reinsert ----
        all_data[f"insert_{rate_key}"] = reinsert_vectors(index, rate_key)

        # Verify count after reinsert
        print("  [WAIT] 5s pause before verify count after reinsert ...")
        time.sleep(5)
        verify_count(index, expected=TOTAL_VECTORS)

        # ---- Bench post-insert ----
        all_data[f"post_insert_{rate_key}"] = run_all_filter_rates(
            f"Post-Reinsert {rate_key}"
        )

        # Wait before next delete (skip after last)
        if rate_key != DELETE_RATES[-1]:
            print(f"\n[WAIT] 10s pause before next delete cycle ...")
            time.sleep(10)

    # ---- Save results ----
    write_excel(all_data, OUTPUT_EXCEL)

    print("\n" + "=" * 65)
    print(" Stability test complete!")
    print(f" Results: {OUTPUT_EXCEL}")
    print("=" * 65)


if __name__ == "__main__":
    main()
