#!/usr/bin/env python3
"""
Rebuild an Excel report directly from already-written vectordbbench result JSON files.

Use this when a benchmark run already completed (all the result_*.json files are sitting
in RESULTS_DIR) but the Excel step failed or was skipped for any reason (e.g. OUTPUT_DIR
was left as a placeholder). This does NOT re-run any benchmarks — it only reads the
existing JSON files and reconstructs the report from them, so it's safe and fast.

It works for both LabelFilterPerformanceCase and NewIntFilterPerformanceCase results, and
for any number of distinct filter_boost_percentage values found in the matched files
(each row gets its own "Boost %" column instead of assuming a single boost).

How matching works:
  Every JSON file has a top-level "task_label" field. vectordbbench sets this to
  "<your --task-label>_<random uuid>" for each run, so files from one script run all
  share the same TASK_LABEL_PREFIX. Set TASK_LABEL_PREFIX below to the --task-label value
  you used (e.g. "30072026") and every file whose task_label starts with that prefix is
  included.

  Caution: the uuid is generated fresh per individual vectordbbench invocation, not once
  per script run — so TASK_LABEL_PREFIX alone can also match files from a *different*
  session that happened to reuse the same --task-label text (e.g. a leftover hardcoded
  default). Set RUN_DATE to the calendar date you actually ran this (matches the real
  write-date vectordbbench stamps into the filename) to additionally scope matching to
  that one day and avoid cross-session collisions.
"""

import glob as glob_module
import json
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ============================================================
# CONFIGURATION
# ============================================================
RESULTS_DIR       = "<PATH_TO_VECTORDB_BENCH_RESULTS_DIR>/Endee"  # <-- set before running (the box where the benchmark actually ran)
TASK_LABEL_PREFIX = "<TASK_LABEL>"               # <-- the --task-label value used in the original run
RUN_DATE          = "<YYYYMMDD>"                 # <-- the calendar date you ran this, e.g. "20260803".
                                                  #     vectordbbench stamps this into the filename itself
                                                  #     (result_<RUN_DATE>_<task_label>_endee.json) using the
                                                  #     real write date, independent of whatever text you chose
                                                  #     for --task-label. Since --task-label defaults get reused
                                                  #     across unrelated sessions/days, TASK_LABEL_PREFIX alone can
                                                  #     match files from a different run that happened to use the
                                                  #     same label text. Combining both narrows it to one specific
                                                  #     day's files. Leave as "" to skip date filtering.
OUTPUT_DIR        = "<PATH_TO_VECTORDB_BENCH_ROOT_DIR>"  # <-- set before running

OUTPUT_EXCEL = os.path.join(OUTPUT_DIR, "rebuilt_bench_from_results.xlsx")


# ============================================================
# LOAD + PARSE RESULT FILES
# ============================================================

def load_matching_results() -> list:
    files = sorted(
        glob_module.glob(os.path.join(RESULTS_DIR, "*_endee.json")),
        key=os.path.getmtime,
    )

    matched = []
    for path in files:
        basename = os.path.basename(path)
        if RUN_DATE and not basename.startswith(f"result_{RUN_DATE}_"):
            continue

        with open(path) as f:
            data = json.load(f)

        if not data.get("task_label", "").startswith(TASK_LABEL_PREFIX):
            continue

        for case_result in data.get("results", []):
            matched.append((path, case_result))

    return matched


def extract_row(path: str, case_result: dict) -> dict:
    task_config    = case_result["task_config"]
    case_config    = task_config["case_config"]
    custom_case    = case_config.get("custom_case") or {}
    db_config      = task_config["db_config"]
    metrics        = case_result["metrics"]

    if "label_percentage" in custom_case:
        filter_type  = "label_pct"
        filter_value = custom_case["label_percentage"]
        case_name    = "LabelFilterPerformanceCase(StrEqual)"
    elif "filter_rate" in custom_case:
        filter_type  = "filter_rate"
        filter_value = custom_case["filter_rate"]
        case_name    = "NewIntFilterPerf"
    else:
        filter_type  = "unknown"
        filter_value = None
        case_name    = str(case_config.get("case_id"))

    concurrency = case_config.get("concurrency_search_config", {}).get("num_concurrency")
    if isinstance(concurrency, list):
        concurrency = ",".join(str(c) for c in concurrency)

    return {
        "mtime":         os.path.getmtime(path),
        "dataset":       custom_case.get("dataset_with_size_type"),
        "precision":     db_config.get("precision"),
        "case_name":     case_name,
        "filter_type":   filter_type,
        "filter_value":  filter_value,
        "prefilter":     db_config.get("prefilter_cardinality_threshold"),
        "boost_pct":     db_config.get("filter_boost_percentage"),
        "m":             db_config.get("m"),
        "ef_search":     db_config.get("ef_search"),
        "ef_con":        db_config.get("ef_con"),
        "top_k":         case_config.get("k"),
        "concurrency":   concurrency,
        "recall":        metrics.get("recall"),
        "qps":           metrics.get("qps"),
        "p99_latency":   metrics.get("serial_latency_p99"),
        "load_duration": metrics.get("load_duration"),
    }


def assign_run_numbers(rows: list) -> list:
    groups = {}
    for r in sorted(rows, key=lambda r: r["mtime"]):
        key = (r["filter_type"], r["filter_value"], r["prefilter"], r["boost_pct"])
        groups.setdefault(key, []).append(r)

    for group_rows in groups.values():
        for i, r in enumerate(group_rows, start=1):
            r["attempt"] = i

    return rows


# ============================================================
# EXCEL WRITER
# ============================================================

def write_excel(rows: list, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rebuilt Bench"

    thin   = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")

    HDR_BG   = "2D6A4F"
    HDR_FONT = Font(bold=True, color="FFFFFF")
    ROW_ODD  = "F0F7F4"
    ROW_EVEN = "FFFFFF"

    columns = [
        ("Dataset",        22),
        ("Precision",      12),
        ("Filter Case",    26),
        ("Prefilter Threshold", 16),
        ("Filter Type",    12),
        ("Filter Value",   13),
        ("Boost %",        10),
        ("Run #",           8),
        ("m",               7),
        ("ef_search",      11),
        ("ef_con",         10),
        ("topK",            8),
        ("Concurrency",    14),
        ("Recall",         10),
        ("QPS",            12),
        ("Latency (p99)(in sec)",  16),
        ("Load Duration(in sec)",  16),
    ]

    for col_idx, (_, width) in enumerate(columns, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28
    for col_idx, (header, _) in enumerate(columns, start=1):
        c = ws.cell(row=1, column=col_idx, value=header)
        c.font      = HDR_FONT
        c.fill      = PatternFill("solid", fgColor=HDR_BG)
        c.alignment = center
        c.border    = border

    rows_sorted = sorted(
        rows,
        key=lambda r: (r["boost_pct"], -r["filter_value"], r["prefilter"] or 0, r["attempt"]),
    )

    current_row = 2
    for row_local_idx, r in enumerate(rows_sorted):
        bg = ROW_ODD if row_local_idx % 2 == 0 else ROW_EVEN
        rf = PatternFill("solid", fgColor=bg)
        ws.row_dimensions[current_row].height = 22

        values = [
            r["dataset"],
            r["precision"],
            r["case_name"],
            r["prefilter"],
            r["filter_type"],
            r["filter_value"],
            r["boost_pct"],
            r["attempt"],
            r["m"],
            r["ef_search"],
            r["ef_con"],
            r["top_k"],
            r["concurrency"],
            round(r["recall"] * 100, 2)  if r["recall"]        is not None else "N/A",
            round(r["qps"], 4)           if r["qps"]            is not None else "N/A",
            round(r["p99_latency"], 6)   if r["p99_latency"]    is not None else "N/A",
            round(r["load_duration"], 4) if r["load_duration"]  is not None else "N/A",
        ]

        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=current_row, column=col_idx, value=val)
            c.fill      = rf
            c.alignment = center if col_idx != 1 else left
            c.border    = border

        current_row += 1

    wb.save(output_path)
    print(f"[EXCEL] Saved → {output_path}")


# ============================================================
# MAIN
# ============================================================

def main():
    matched = load_matching_results()
    if not matched:
        print(
            f"[ERROR] No result files under {RESULTS_DIR} with task_label starting with "
            f"'{TASK_LABEL_PREFIX}'" + (f" on date {RUN_DATE}" if RUN_DATE else "")
        )
        return

    print(f"[INFO] Found {len(matched)} matching result(s)")
    rows = [extract_row(path, case_result) for path, case_result in matched]
    rows = assign_run_numbers(rows)
    write_excel(rows, OUTPUT_EXCEL)


if __name__ == "__main__":
    main()
