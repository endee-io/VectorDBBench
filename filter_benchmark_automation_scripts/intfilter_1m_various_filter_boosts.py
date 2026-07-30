#!/usr/bin/env python3
"""
Int Filter Benchmark Script - NewIntFilterPerformanceCase
Collection: (set COLLECTION_NAME below — 1M vectors pre-loaded, no load step)

Workflow:
  For each filter_boost_percentage in [0, 25, 50, 75, 100]:
    Run vectordbbench for filter rates: 0.01, 0.50, 0.80, 0.99
    For filter_rate=0.99: also run with prefilter_cardinality_threshold=50000
    10-second gap between every run.

Output: Excel file with one table per boost percentage.
"""

import subprocess
import json
import os
import time
import glob as glob_module
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# ============================================================
# CONFIGURATION
# ============================================================
TOKEN             = "<ENDEE_API_TOKEN>"          # <-- set before running
BASE_URL          = "http://localhost:8080/api/v2"  # <-- set before running
COLLECTION_NAME   = "<COLLECTION_NAME>"          # <-- set before running (must already exist, pre-loaded)
DATASET_LOCAL_DIR = "<PATH_TO_LOCAL_DATASET_DIR>"  # <-- set before running
RESULTS_DIR       = "<PATH_TO_VECTORDB_BENCH_RESULTS_DIR>/Endee"  # <-- set before running
REGION            = "<ENDEE_REGION>"             # <-- set before running
TASK_LABEL        = "<TASK_LABEL>"               # <-- set before running (e.g. a date or run identifier)

M               = 16
EF_SEARCH       = 128
EF_CON          = 128
TOP_K           = 30
CONCURRENCY     = 16
CONCURRENCY_DUR = 30
PRECISION       = "int16"

FILTER_RATES      = [0.01, 0.50, 0.80, 0.99]
PRECARDINALITY    = 50000          # extra run for filter_rate=0.99
DEFAULT_PRECARDINALITY = 10000     # Endee's backend default prefilter_cardinality_threshold when the flag isn't passed
BOOST_PERCENTAGES = [0, 25, 50, 75, 100]

RUNS_PER_CONFIG   = 3               # repeat each config this many times, keep the best QPS

OUTPUT_DIR = "<PATH_TO_VECTORDB_BENCH_ROOT_DIR>"  # <-- set before running

OUTPUT_EXCEL = os.path.join(
    OUTPUT_DIR,
    f"intfilter_bench_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
)

# ============================================================
# BENCHMARK RUNNER
# ============================================================

def run_vectordbbench_once(filter_rate: float, boost_pct: int, prefilter: int = None, attempt: int = 1) -> dict:
    before = set(glob_module.glob(os.path.join(RESULTS_DIR, "*.json")))

    prefilter_part = f"--prefilter-cardinality-threshold {prefilter} " if prefilter else ""
    boost_part     = f"--filter-boost-percentage {boost_pct} " if boost_pct is not None else ""

    cmd = (
        f'DATASET_LOCAL_DIR="{DATASET_LOCAL_DIR}" vectordbbench endee '
        f'--token "{TOKEN}" '
        f'--region {REGION} '
        f'--base-url "{BASE_URL}" '
        f'--collection-name {COLLECTION_NAME} '
        f'--task-label "{TASK_LABEL}" '
        f'--m {M} '
        f'--ef-con {EF_CON} '
        f'--ef-search {EF_SEARCH} '
        f'--space-type cosine '
        f'{prefilter_part}'
        f'{boost_part}'
        f'--precision {PRECISION} '
        f'--version 1 '
        f'--case-type NewIntFilterPerformanceCase '
        f'--dataset-with-size-type "Medium Cohere (768dim, 1M)" '
        f'--filter-rate {filter_rate} '
        f'--k {TOP_K} '
        f'--num-concurrency "{CONCURRENCY}" '
        f'--concurrency-duration {CONCURRENCY_DUR} '
        f'--concurrency-timeout 3600 '
        f'--skip-drop-old '
        f'--skip-load '
        f'--search-concurrent '
        f'--search-serial'
    )

    label = f"filter_rate={filter_rate}, boost={boost_pct}%" + (f", prefilter={prefilter}" if prefilter else "")
    print(f"\n  [RUN attempt {attempt}/{RUNS_PER_CONFIG}] {label}")
    proc = subprocess.run(cmd, shell=True, text=True)
    if proc.returncode != 0:
        print(f"  [WARN] vectordbbench exited with code {proc.returncode}")

    time.sleep(5)
    after = set(glob_module.glob(os.path.join(RESULTS_DIR, "*.json")))
    new_files = after - before

    if not new_files:
        print(f"  [ERROR] No new result file for {label}")
        return {"recall": None, "qps": None, "p99_latency": None, "load_duration": None}

    result_file = max(new_files, key=os.path.getmtime)
    print(f"  [FILE] {os.path.basename(result_file)}")

    with open(result_file) as f:
        data = json.load(f)

    metrics  = data["results"][0]["metrics"]
    recall   = metrics.get("recall")
    qps      = metrics.get("qps")
    p99      = metrics.get("serial_latency_p99")
    load_dur = metrics.get("load_duration")

    print(f"  [METRICS] recall={recall}, qps={qps}, p99={p99}, load_duration={load_dur}")
    return {"recall": recall, "qps": qps, "p99_latency": p99, "load_duration": load_dur}


def run_vectordbbench(filter_rate: float, boost_pct: int, prefilter: int = None) -> dict:
    """Run the same config RUNS_PER_CONFIG times and return the result with the best (highest) QPS."""
    best = None
    for attempt in range(1, RUNS_PER_CONFIG + 1):
        metrics = run_vectordbbench_once(filter_rate, boost_pct, prefilter, attempt)
        if best is None or (metrics["qps"] is not None and (best["qps"] is None or metrics["qps"] > best["qps"])):
            best = metrics

        if attempt < RUNS_PER_CONFIG:
            print(f"  [WAIT] 10s before repeat run ...")
            time.sleep(10)

    print(f"  [BEST OF {RUNS_PER_CONFIG}] qps={best['qps']}")
    return best


# ============================================================
# EXCEL WRITER
# ============================================================

def write_excel(all_data: dict, output_path: str):
    """
    all_data: {boost_pct: [row_dict, ...]}
    Each row_dict has: filter_rate, prefilter, recall, qps, p99_latency, load_duration
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Int Filter Bench"

    thin   = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")

    HDR_BG   = "2D6A4F"
    HDR_FONT = Font(bold=True, color="FFFFFF")
    ROW_ODD  = "F0F7F4"
    ROW_EVEN = "FFFFFF"

    DATASET_NAME = "Cohere 1M (768D)"
    FILTER_CASE  = "NewIntFilterPerf"

    columns = [
        ("Dataset",        22),
        ("Precision",      12),
        ("Filter Case",    22),
        ("Filter Rate",    13),
        ("m",               7),
        ("ef_search",      11),
        ("ef_con",         10),
        ("topK",            8),
        ("Concurrency",    14),
        ("Recall",         10),
        ("QPS",            12),
        ("Latency (p99)(in sec)",  16),
        ("Load Duration(in sec)",  16),
    ]

    NUM_COLS = len(columns)

    # Set column widths once
    for col_idx, (_, width) in enumerate(columns, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    current_row = 1

    for boost_pct in BOOST_PERCENTAGES:
        rows = all_data.get(boost_pct, [])

        # --- Boost label row (outside table) ---
        label_cell = ws.cell(row=current_row, column=1,
                             value=f"Filter Boost Percentage: {boost_pct}%")
        label_cell.font      = Font(bold=True, size=12)
        label_cell.alignment = left
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row,   end_column=NUM_COLS
        )
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # --- Header row ---
        ws.row_dimensions[current_row].height = 28
        for col_idx, (header, _) in enumerate(columns, start=1):
            c = ws.cell(row=current_row, column=col_idx, value=header)
            c.font      = HDR_FONT
            c.fill      = PatternFill("solid", fgColor=HDR_BG)
            c.alignment = center
            c.border    = border
        current_row += 1

        # --- Data rows ---
        for row_local_idx, r in enumerate(rows):
            ws.row_dimensions[current_row].height = 22
            bg = ROW_ODD if row_local_idx % 2 == 0 else ROW_EVEN
            rf = PatternFill("solid", fgColor=bg)

            case_label = FILTER_CASE
            case_label += f"\n(prefilter={r['prefilter'] or DEFAULT_PRECARDINALITY})"

            recall   = r.get("recall")
            qps      = r.get("qps")
            p99      = r.get("p99_latency")
            load_dur = r.get("load_duration")

            values = [
                DATASET_NAME,
                PRECISION,
                case_label,
                r["filter_rate"],
                M,
                EF_SEARCH,
                EF_CON,
                TOP_K,
                CONCURRENCY,
                round(recall * 100, 2)  if recall   is not None else "N/A",
                round(qps, 4)           if qps       is not None else "N/A",
                round(p99, 6)           if p99       is not None else "N/A",
                round(load_dur, 4)      if load_dur  is not None else "N/A",
            ]

            for col_idx, val in enumerate(values, start=1):
                c = ws.cell(row=current_row, column=col_idx, value=val)
                c.fill      = rf
                c.alignment = center if col_idx != 1 else left
                c.border    = border

            current_row += 1

        # 2 blank rows before next section
        current_row += 2

    wb.save(output_path)
    print(f"\n[EXCEL] Saved → {output_path}")


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 60)
    print("Int Filter Benchmark")
    print(f"Collection : {COLLECTION_NAME}")
    print(f"Output: {OUTPUT_EXCEL}")
    print("=" * 60)

    all_data = {}

    for boost_pct in BOOST_PERCENTAGES:
        print(f"\n{'='*60}")
        print(f"BOOST PERCENTAGE: {boost_pct}%")
        print(f"{'='*60}")
        rows = []

        for fr in FILTER_RATES:
            metrics = run_vectordbbench(fr, boost_pct)
            rows.append({
                "filter_rate": fr,
                "prefilter":   None,
                **metrics,
            })

            if fr == 0.99:
                print(f"\n  [WAIT] 10s before precardinality run ...")
                time.sleep(10)
                metrics_pre = run_vectordbbench(fr, boost_pct, prefilter=PRECARDINALITY)
                rows.append({
                    "filter_rate": fr,
                    "prefilter":   PRECARDINALITY,
                    **metrics_pre,
                })

            print(f"\n  [WAIT] 10s before next run ...")
            time.sleep(10)

        all_data[boost_pct] = rows

    write_excel(all_data, OUTPUT_EXCEL)


if __name__ == "__main__":
    main()