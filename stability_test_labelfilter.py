#!/usr/bin/env python3
"""
Stability Test Script - Label Filter
Index: test_1M_vaib_filter_1503_adup2 (1M vectors pre-loaded, no load step)

Workflow:
  1. Fresh run     -> 8 label percentages (0.001, 0.002, 0.01, 0.02, 0.05, 0.1, 0.2, 0.5)
  For each label in [label_1p, label_2p, label_5p, label_10p, label_20p, label_50p]:
  2. Delete <label>  -> verify count -> bench 8 label percentages
  3. Wait 10s        -> Reinsert <label> -> verify count -> bench 8 label percentages
  4. Wait 10s        -> next label

Recorded per bench run: Recall (%), QPS, P99 Latency (s)
Recorded per operation: Delete duration (s), Insert duration (s)
Output: Excel file with all results
"""

import subprocess
import json
import os
import time
import glob as glob_module
import pyarrow.parquet as pq
from endee import Endee as EndeeClient
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# ============================================================
# CONFIGURATION
# ============================================================
TOKEN = "localtest"
BASE_URL = "http://148.113.54.173:8080/api/v1"
INDEX_NAME = "test_1M_vaib_filter_1503_adup2"
DATASET_LOCAL_DIR = "/home/debian/latest_VDB/VectorDBBench/vectordataset_label"
BASE_PATH = "/home/debian/latest_VDB/VectorDBBench/vectordataset_label"
DATASET_FOLDER = "cohere/cohere_medium_1m"
RESULTS_DIR = "/home/debian/latest_VDB/VectorDBBench/vectordb_bench/results/Endee"

LABEL_PERCENTAGES = [0.001, 0.002, 0.01, 0.02, 0.05, 0.10, 0.20, 0.50]

# Labels to delete/reinsert in order
DELETE_LABELS = ["label_1p", "label_2p", "label_5p", "label_10p", "label_20p", "label_50p"]

# Expected count after each label is deleted (non-cumulative — each cycle starts from 1M)
EXPECTED_COUNT_AFTER_DELETE = {
    "label_1p":  990000,
    "label_2p":  980000,
    "label_5p":  950000,
    "label_10p": 900000,
    "label_20p": 800000,
    "label_50p": 500000,
}

BATCH_SIZE = 1000

OUTPUT_EXCEL = os.path.join(
    "/home/debian/latest_VDB/VectorDBBench",
    f"stability_test_labelfilter_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
)

# ============================================================
# VECTORDBBENCH RUNNER
# ============================================================

def run_vectordbbench(label_pct: float) -> dict:
    """Run vectordbbench for one label percentage; return parsed metrics dict."""
    before = set(glob_module.glob(os.path.join(RESULTS_DIR, "*.json")))

    cmd = (
        f'DATASET_LOCAL_DIR="{DATASET_LOCAL_DIR}" vectordbbench endee '
        f'--token "{TOKEN}" '
        f'--region india-west-1 '
        f'--base-url "{BASE_URL}" '
        f'--index-name {INDEX_NAME} '
        f'--task-label "20260107" '
        f'--m 16 '
        f'--ef-con 128 '
        f'--ef-search 128 '
        f'--space-type cosine '
        f'--precision int16 '
        f'--version 1 '
        f'--case-type LabelFilterPerformanceCase '
        f'--dataset-with-size-type "Medium Cohere (768dim, 1M)" '
        f'--label-percentage {label_pct} '
        f'--k 30 '
        f'--num-concurrency "16" '
        f'--concurrency-duration 30 '
        f'--concurrency-timeout 3600 '
        f'--skip-drop-old '
        f'--skip-load '
        f'--search-concurrent '
        f'--search-serial'
    )

    print(f"\n  [RUN] label_percentage={label_pct}")
    proc = subprocess.run(cmd, shell=True, text=True)
    if proc.returncode != 0:
        print(f"  [WARN] vectordbbench exited with code {proc.returncode}")

    time.sleep(3)
    after = set(glob_module.glob(os.path.join(RESULTS_DIR, "*.json")))
    new_files = after - before

    if not new_files:
        print(f"  [ERROR] No new result file found for label_percentage={label_pct}")
        return {"recall": None, "qps": None, "p99_latency": None}

    result_file = max(new_files, key=os.path.getmtime)
    print(f"  [FILE] {os.path.basename(result_file)}")

    with open(result_file) as f:
        data = json.load(f)

    metrics = data["results"][0]["metrics"]
    recall = metrics.get("recall")
    qps    = metrics.get("qps")
    p99    = metrics.get("serial_latency_p99")

    print(f"  [METRICS] recall={recall}, qps={qps}, p99={p99}")
    return {"recall": recall, "qps": qps, "p99_latency": p99}


def run_all_label_percentages(phase_label: str) -> dict:
    """Run vectordbbench for all 8 label percentages. Returns {label_pct: metrics}."""
    print(f"\n==> Running all label percentages [{phase_label}]")
    results = {}
    for lp in LABEL_PERCENTAGES:
        results[lp] = run_vectordbbench(lp)
        print("  [WAIT] 5s gap before next label percentage run ...")
        time.sleep(5)
    return results


# ============================================================
# DELETE / VERIFY / REINSERT
# ============================================================

def delete_vectors(index, label: str) -> float:
    """Delete vectors matching label. Returns elapsed seconds."""
    print(f"\n==> DELETE {label}: removing vectors with label == '{label}'")
    start = time.perf_counter()
    result = index.delete_with_filter([{"label": {"$eq": label}}])
    elapsed = time.perf_counter() - start
    print(f"  [DELETE] Completed in {elapsed:.2f}s | result: {result}")
    print("  [WAIT] 5s pause before verify count ...")
    time.sleep(5)
    return elapsed


def verify_count(index, expected: int, context: str) -> int:
    """Call index.describe() and print count vs expected."""
    desc = index.describe()
    actual = desc.get("count", "UNKNOWN")
    status = "OK" if actual == expected else "MISMATCH"
    print(f"  [VERIFY:{context}] Expected={expected}  Actual={actual}  [{status}]")
    return actual


def reinsert_vectors(index, label: str) -> float:
    """Reinsert vectors matching label from parquet. Returns elapsed seconds."""
    dataset_path = os.path.join(BASE_PATH, DATASET_FOLDER)
    labels_path  = os.path.join(dataset_path, "scalar_labels.parquet")
    emb_path     = os.path.join(dataset_path, "shuffle_train.parquet")

    print(f"\n==> REINSERT {label}: reading scalar_labels.parquet ...")
    labels_df = pq.ParquetFile(labels_path).read().to_pandas()
    filtered_labels = labels_df[labels_df["labels"] == label]
    valid_ids = set(filtered_labels["id"].values)
    print(f"  [INSERT] Total vectors to reinsert: {len(valid_ids)}")

    emb_file = pq.ParquetFile(emb_path)
    total = 0
    start = time.perf_counter()

    for batch in emb_file.iter_batches(batch_size=BATCH_SIZE):
        batch_df = batch.to_pandas()
        batch_df = batch_df[batch_df["id"].isin(valid_ids)]
        if batch_df.empty:
            continue

        batch_df = batch_df.merge(filtered_labels[["id", "labels"]], on="id")

        vectors = [
            {
                "id": str(row["id"]),
                "vector": row["emb"].tolist() if hasattr(row["emb"], "tolist") else list(row["emb"]),
                "meta": {"id": row["id"]},
                "filter": {"id": row["id"], "label": row["labels"]},
            }
            for _, row in batch_df.iterrows()
        ]
        index.upsert(vectors)
        total += len(vectors)
        print(f"  [INSERT] upserted {len(vectors)}, cumulative: {total}")

    elapsed = time.perf_counter() - start
    print(f"  [INSERT] Done. Total={total} vectors in {elapsed:.2f}s")
    return elapsed


# ============================================================
# EXCEL WRITER
# ============================================================

def write_excel(all_data: dict, output_path: str):
    """Write all collected results to a formatted Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stability Test LabelFilter"

    def hdr(bold=True, color="FFFFFF"):
        return Font(bold=bold, color=color)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")
    thin   = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    COLOR_DARK   = "2E4057"
    COLOR_BLUE   = "1565C0"
    COLOR_LBLUE  = "BBDEFB"
    COLOR_GREEN  = "1B5E20"
    COLOR_LGREEN = "C8E6C9"
    COLOR_ODD    = "F5F5F5"

    # Number of filter rates
    n_fr = len(LABEL_PERCENTAGES)
    # Col 1 = Phase, Cols 2..(1+n_fr*3) = filter rate groups, last 2 = durations
    dur_col_start = 2 + n_fr * 3

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 28

    # Phase header (merge rows 1-2)
    ws.merge_cells("A1:A2")
    c = ws["A1"]
    c.value = "Phase"
    c.font = hdr(); c.fill = fill(COLOR_DARK)
    c.alignment = center; c.border = border

    # Filter rate group headers (row 1)
    lp_labels = {lp: f"Label Pct {lp}" for lp in LABEL_PERCENTAGES}
    col = 2
    for lp in LABEL_PERCENTAGES:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
        c = ws.cell(row=1, column=col, value=lp_labels[lp])
        c.font = hdr(); c.fill = fill(COLOR_BLUE)
        c.alignment = center; c.border = border
        col += 3

    # Operation duration header (row 1)
    ws.merge_cells(start_row=1, start_column=dur_col_start, end_row=1, end_column=dur_col_start + 1)
    c = ws.cell(row=1, column=dur_col_start, value="Operation Duration")
    c.font = hdr(); c.fill = fill(COLOR_GREEN)
    c.alignment = center; c.border = border

    # Sub-headers (row 2)
    col = 2
    for _ in LABEL_PERCENTAGES:
        for sub in ("Recall (%)", "QPS", "P99 Latency (s)"):
            c = ws.cell(row=2, column=col, value=sub)
            c.font = Font(bold=True, color="000000")
            c.fill = fill(COLOR_LBLUE)
            c.alignment = center; c.border = border
            col += 1

    for sub in ("Delete Duration (s)", "Insert Duration (s)"):
        c = ws.cell(row=2, column=col, value=sub)
        c.font = Font(bold=True, color="000000")
        c.fill = fill(COLOR_LGREEN)
        c.alignment = center; c.border = border
        col += 1

    # Build phase rows
    phases = [("Fresh Insertion\n(1M vectors)", "fresh", None, None)]
    for lbl in DELETE_LABELS:
        phases.append((f"After Delete {lbl}", f"post_delete_{lbl}", f"delete_{lbl}", None))
        phases.append((f"After Reinsert {lbl}", f"post_insert_{lbl}", None, f"insert_{lbl}"))

    for i, (phase_name, bench_key, del_key, ins_key) in enumerate(phases):
        row = 3 + i
        ws.row_dimensions[row].height = 32
        row_fill = fill(COLOR_ODD) if i % 2 == 0 else fill("FFFFFF")

        c = ws.cell(row=row, column=1, value=phase_name)
        c.font = Font(bold=True); c.fill = row_fill
        c.alignment = left; c.border = border

        bench_results = all_data.get(bench_key) or {}
        col = 2
        for lp in LABEL_PERCENTAGES:
            m = bench_results.get(lp) or {}
            recall = m.get("recall")
            qps    = m.get("qps")
            p99    = m.get("p99_latency")
            vals = [
                round(recall * 100, 2) if recall is not None else "N/A",
                round(qps, 4)          if qps     is not None else "N/A",
                round(p99, 6)          if p99      is not None else "N/A",
            ]
            for v in vals:
                c = ws.cell(row=row, column=col, value=v)
                c.fill = row_fill; c.alignment = center; c.border = border
                col += 1

        del_dur = all_data.get(del_key) if del_key else None
        ins_dur = all_data.get(ins_key) if ins_key else None

        c = ws.cell(row=row, column=col,
                    value=round(del_dur, 2) if del_dur is not None else "")
        c.fill = row_fill; c.alignment = center; c.border = border

        c = ws.cell(row=row, column=col + 1,
                    value=round(ins_dur, 2) if ins_dur is not None else "")
        c.fill = row_fill; c.alignment = center; c.border = border

    # Column widths
    from openpyxl.utils import get_column_letter
    ws.column_dimensions["A"].width = 30
    for col_idx in range(2, dur_col_start + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    ws.freeze_panes = "B3"
    wb.save(output_path)
    print(f"\n[EXCEL] Saved: {output_path}")


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 65)
    print(" Stability Test — Label Filter")
    print(f" Index : {INDEX_NAME}")
    print(f" Output: {OUTPUT_EXCEL}")
    print("=" * 65)

    client = EndeeClient(token=TOKEN)
    client.set_base_url(BASE_URL)
    index = client.get_index(INDEX_NAME)
    print(f"\n[INIT] Connected to index '{INDEX_NAME}'")
    print(f"[INIT] Current state: {index.describe()}")

    all_data: dict = {}

    # ---- Phase 1: Fresh run ----
    all_data["fresh"] = run_all_label_percentages("Fresh — 1M vectors")

    for label in DELETE_LABELS:
        # ---- Delete ----
        all_data[f"delete_{label}"] = delete_vectors(index, label)

        # Verify count after delete
        verify_count(index, EXPECTED_COUNT_AFTER_DELETE[label], f"post-delete {label}")

        # ---- Bench post-delete ----
        all_data[f"post_delete_{label}"] = run_all_label_percentages(f"Post-Delete {label}")

        # ---- Wait then reinsert ----
        print(f"\n[WAIT] 10s pause before reinsert {label} ...")
        time.sleep(10)

        all_data[f"insert_{label}"] = reinsert_vectors(index, label)

        # Verify count after reinsert (5s wait first)
        print("  [WAIT] 5s pause before verify count after reinsert ...")
        time.sleep(5)
        verify_count(index, 1000000, f"post-reinsert {label}")

        # ---- Bench post-reinsert ----
        all_data[f"post_insert_{label}"] = run_all_label_percentages(f"Post-Reinsert {label}")

        # Wait before next label (skip after last)
        if label != DELETE_LABELS[-1]:
            print(f"\n[WAIT] 10s pause before next delete cycle ...")
            time.sleep(10)

    # ---- Save results ----
    write_excel(all_data, OUTPUT_EXCEL)

    print("\n" + "=" * 65)
    print(" Stability test complete!")
    print(f" Results: {OUTPUT_EXCEL}")
    print("=" * 65)


if __name__ == "__main__":
    main()
