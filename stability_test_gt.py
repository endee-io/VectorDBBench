#!/usr/bin/env python3
"""
Stability Test Script - GT Filter Operator
Index: (set INDEX_NAME below — 1M vectors pre-loaded, no load step)

Dataset: vectordataset_gt
Delete filter: $lte  (e.g. delete id <= 10000 removes 10001 vectors for 1p)

Workflow:
  1. Fresh run     -> 4 filter rates (0.99, 0.80, 0.50, 0.01)
                      + 0.99 re-run with prefilter_cardinality_threshold=9999
  2. Delete 1p     -> verify count (989999) -> bench 4 rates + 0.99 prefilter
  3. Wait 10s      -> Reinsert 1p  -> verify count (1000000) -> bench 4 rates + 0.99 prefilter
  4. Wait 10s      -> Delete 50p   -> verify count (499999)  -> bench 4 rates + 0.99 prefilter
  5. Wait 10s      -> Reinsert 50p -> verify count (1000000) -> bench 4 rates + 0.99 prefilter
  6. Wait 10s      -> Delete 80p   -> verify count (199999)  -> bench 4 rates + 0.99 prefilter
  7. Wait 10s      -> Reinsert 80p -> verify count (1000000) -> bench 4 rates + 0.99 prefilter

Delete counts note:
  $lte 10000  -> ids 0..10000 inclusive = 10001 deleted -> 989999 remain
  $lte 500000 -> 500001 deleted -> 499999 remain
  $lte 800000 -> 800001 deleted -> 199999 remain

Recorded per bench run : Recall (%), QPS, P99 Latency (s)
Recorded per operation : Delete duration (s), Insert duration (s)
Output                 : Excel file with all results
"""

import subprocess
import json
import os
import time
import glob as glob_module
import pyarrow.parquet as pq
from endee import Endee as EndeeClient
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# ============================================================
# CONFIGURATION
# ============================================================
TOKEN             = "localtest"
BASE_URL          = "http://148.113.54.173:8080/api/v1"
INDEX_NAME        = "test_1M_vaib_filter_1603_adup5"   # <-- set before running
DATASET_LOCAL_DIR = "/home/debian/latest_VDB/VectorDBBench/vectordataset_gt"
BASE_PATH         = "/home/debian/latest_VDB/VectorDBBench/vectordataset_gt"
DATASET_FOLDER    = "cohere/cohere_medium_1m"
RESULTS_DIR       = "/home/debian/latest_VDB/VectorDBBench/vectordb_bench/results/Endee"

FILTER_RATES = [0.99, 0.80, 0.50, 0.01]
DELETE_RATES = ["1p", "50p", "80p"]

# Delete: remove vectors with id <= threshold (includes threshold id)
RANGE_MAP_DELETE = {
    "1p":  10000,    # $lte 10000 -> 10001 deleted -> 989999 remain
    "50p": 500000,   # $lte 500000 -> 500001 deleted -> 499999 remain
    "80p": 800000,   # $lte 800000 -> 800001 deleted -> 199999 remain
}

# Reinsert: re-upsert vectors in [low, high] inclusive
RANGE_MAP_INSERT = {
    "1p":  [0, 10000],
    "50p": [0, 500000],
    "80p": [0, 800000],
}

# Expected count after delete
EXPECTED_COUNT_AFTER_DELETE = {
    "1p":  989999,
    "50p": 499999,
    "80p": 199999,
}

TOTAL_VECTORS = 1000000
BATCH_SIZE    = 1000

OUTPUT_EXCEL = os.path.join(
    "/home/debian/latest_VDB/VectorDBBench",
    f"stability_test_gt_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
)

# ============================================================
# VECTORDBBENCH RUNNER
# ============================================================

def run_vectordbbench(filter_rate: float, prefilter: int | None = None) -> dict:
    """Run vectordbbench for one filter rate; return parsed metrics dict."""
    before = set(glob_module.glob(os.path.join(RESULTS_DIR, "*.json")))

    prefilter_flag = f'--prefilter-cardinality-threshold {prefilter} ' if prefilter else ''

    cmd = (
        f'DATASET_LOCAL_DIR="{DATASET_LOCAL_DIR}" vectordbbench endee '
        f'--token "{TOKEN}" '
        f'--region india-west-1 '
        f'--base-url "{BASE_URL}" '
        f'--index-name {INDEX_NAME} '
        f'--task-label "20260107" '
        f'--m 16 '
        f'--ef-con 128 '
        f'--ef-search 128 '
        f'--space-type cosine '
        f'{prefilter_flag}'
        f'--precision int16 '
        f'--version 1 '
        f'--case-type NewIntFilterPerformanceCase '
        f'--dataset-with-size-type "Medium Cohere (768dim, 1M)" '
        f'--filter-rate {filter_rate} '
        f'--k 30 '
        f'--num-concurrency "16" '
        f'--concurrency-duration 30 '
        f'--concurrency-timeout 3600 '
        f'--skip-drop-old '
        f'--skip-load '
        f'--search-concurrent '
        f'--search-serial'
    )

    prefilter_info = f", prefilter={prefilter}" if prefilter else ""
    print(f"\n  [RUN] filter_rate={filter_rate}{prefilter_info}")
    proc = subprocess.run(cmd, shell=True, text=True)
    if proc.returncode != 0:
        print(f"  [WARN] vectordbbench exited with code {proc.returncode}")

    time.sleep(3)
    after = set(glob_module.glob(os.path.join(RESULTS_DIR, "*.json")))
    new_files = after - before

    if not new_files:
        print(f"  [ERROR] No new result file found for filter_rate={filter_rate}{prefilter_info}")
        return {"recall": None, "qps": None, "p99_latency": None}

    result_file = max(new_files, key=os.path.getmtime)
    print(f"  [FILE] {os.path.basename(result_file)}")

    with open(result_file) as f:
        data = json.load(f)

    metrics = data["results"][0]["metrics"]
    recall = metrics.get("recall")
    qps    = metrics.get("qps")
    p99    = metrics.get("serial_latency_p99")

    print(f"  [METRICS] recall={recall}, qps={qps}, p99={p99}")
    return {"recall": recall, "qps": qps, "p99_latency": p99}


def run_all_filter_rates(label: str) -> dict:
    """
    Run vectordbbench for all 4 filter rates.
    For filter_rate=0.99, also runs a second time with prefilter=9999.
    Returns {filter_rate: metrics, "0.99_prefilter": metrics}.
    """
    print(f"\n==> Running all filter rates [{label}]")
    results = {}
    for fr in FILTER_RATES:
        results[fr] = run_vectordbbench(fr)
        print("  [WAIT] 5s gap before next filter rate run ...")
        time.sleep(5)
        if fr == 0.99:
            print(f"  [RUN] filter_rate=0.99 (prefilter=9999) ...")
            results["0.99_prefilter"] = run_vectordbbench(0.99, prefilter=9999)
            print("  [WAIT] 5s gap before next filter rate run ...")
            time.sleep(5)
    return results


# ============================================================
# DELETE / VERIFY / REINSERT
# ============================================================

def delete_vectors(index, rate_key: str) -> float:
    """Delete vectors with id <= threshold ($lte). Returns elapsed seconds."""
    threshold = RANGE_MAP_DELETE[rate_key]
    print(f"\n==> DELETE {rate_key}: removing vectors with id <= {threshold}")
    start = time.perf_counter()
    result = index.delete_with_filter([{"id": {"$lte": threshold}}])
    elapsed = time.perf_counter() - start
    print(f"  [DELETE] Completed in {elapsed:.2f}s | result: {result}")
    print("  [WAIT] 5s pause before verify count ...")
    time.sleep(5)
    return elapsed


def verify_count(index, expected: int, context: str = "") -> int:
    """Call index.describe() and print count vs expected."""
    desc = index.describe()
    actual = desc.get("count", "UNKNOWN")
    status = "OK" if actual == expected else "MISMATCH"
    tag = f" [{context}]" if context else ""
    print(f"  [VERIFY{tag}] Expected={expected}  Actual={actual}  [{status}]")
    return actual


def reinsert_vectors(index, rate_key: str) -> float:
    """Reinsert vectors in [0, threshold]. Returns elapsed seconds."""
    low, high = RANGE_MAP_INSERT[rate_key]
    emb_path = os.path.join(BASE_PATH, DATASET_FOLDER, "shuffle_train.parquet")
    print(f"\n==> REINSERT {rate_key}: upserting vectors with id in [{low}, {high}]")

    emb_file = pq.ParquetFile(emb_path)
    total = 0
    start = time.perf_counter()

    for batch in emb_file.iter_batches(batch_size=BATCH_SIZE):
        df = batch.to_pandas()
        df = df[(df["id"] >= low) & (df["id"] <= high)]
        if df.empty:
            continue

        vectors = [
            {
                "id": str(row["id"]),
                "vector": row["emb"].tolist() if hasattr(row["emb"], "tolist") else list(row["emb"]),
                "meta":   {"id": row["id"]},
                "filter": {"id": row["id"]},
            }
            for _, row in df.iterrows()
        ]
        index.upsert(vectors)
        total += len(vectors)
        print(f"  [INSERT] upserted {len(vectors)}, cumulative: {total}")

    elapsed = time.perf_counter() - start
    print(f"  [INSERT] Done. Total={total} vectors in {elapsed:.2f}s")
    return elapsed


# ============================================================
# EXCEL WRITER
# ============================================================

def write_excel(all_data: dict, output_path: str):
    """Write all collected results to a formatted Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stability Test GT"

    def hdr(bold=True, color="FFFFFF"):
        return Font(bold=bold, color=color)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")
    thin   = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    COLOR_DARK    = "2E4057"
    COLOR_BLUE    = "1565C0"
    COLOR_PURPLE  = "6A1B9A"   # distinct colour for 0.99 prefilter group
    COLOR_LBLUE   = "BBDEFB"
    COLOR_LPURPLE = "E1BEE7"
    COLOR_GREEN   = "1B5E20"
    COLOR_LGREEN  = "C8E6C9"
    COLOR_ROW_ODD = "F5F5F5"

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 28

    # --- Row 1: top-level headers ---
    # Column layout:
    #   A           = Phase
    #   B-D         = Filter Rate 0.99 (normal)
    #   E-G         = Filter Rate 0.99 (prefilter=9999)
    #   H-J         = Filter Rate 0.80
    #   K-M         = Filter Rate 0.50
    #   N-P         = Filter Rate 0.01
    #   Q-R         = Operation Duration

    ws.merge_cells("A1:A2")
    c = ws["A1"]
    c.value = "Phase"
    c.font = hdr(); c.fill = fill(COLOR_DARK)
    c.alignment = center; c.border = border

    # 0.99 normal
    ws.merge_cells("B1:D1")
    c = ws["B1"]
    c.value = "Filter Rate 0.99"
    c.font = hdr(); c.fill = fill(COLOR_BLUE)
    c.alignment = center; c.border = border

    # 0.99 prefilter
    ws.merge_cells("E1:G1")
    c = ws["E1"]
    c.value = "Filter Rate 0.99\n(prefilter_cardinality=9999)"
    c.font = hdr(); c.fill = fill(COLOR_PURPLE)
    c.alignment = center; c.border = border

    # 0.80, 0.50, 0.01
    remaining_rates = [0.80, 0.50, 0.01]
    fr_labels = {0.80: "Filter Rate 0.80", 0.50: "Filter Rate 0.50", 0.01: "Filter Rate 0.01"}
    col = 8  # column H
    for fr in remaining_rates:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
        c = ws.cell(row=1, column=col, value=fr_labels[fr])
        c.font = hdr(); c.fill = fill(COLOR_BLUE)
        c.alignment = center; c.border = border
        col += 3

    # Operation duration (cols Q-R = 17-18)
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
    c = ws.cell(row=1, column=col, value="Operation Duration")
    c.font = hdr(); c.fill = fill(COLOR_GREEN)
    c.alignment = center; c.border = border

    # --- Row 2: sub-headers ---
    sub_cols = [
        ("B", COLOR_LBLUE), ("C", COLOR_LBLUE), ("D", COLOR_LBLUE),   # 0.99 normal
        ("E", COLOR_LPURPLE), ("F", COLOR_LPURPLE), ("G", COLOR_LPURPLE),  # 0.99 prefilter
        ("H", COLOR_LBLUE), ("I", COLOR_LBLUE), ("J", COLOR_LBLUE),   # 0.80
        ("K", COLOR_LBLUE), ("L", COLOR_LBLUE), ("M", COLOR_LBLUE),   # 0.50
        ("N", COLOR_LBLUE), ("O", COLOR_LBLUE), ("P", COLOR_LBLUE),   # 0.01
    ]
    for col_letter, bg in sub_cols:
        c = ws[f"{col_letter}2"]
        # cycle Recall/QPS/P99 by position
        idx = "BCDEFGHIJKLMNOP".index(col_letter)
        c.value = ("Recall (%)", "QPS", "P99 Latency (s)")[idx % 3]
        c.font = Font(bold=True, color="000000")
        c.fill = fill(bg)
        c.alignment = center; c.border = border

    for col_letter, label in [("Q", "Delete Duration (s)"), ("R", "Insert Duration (s)")]:
        c = ws[f"{col_letter}2"]
        c.value = label
        c.font = Font(bold=True, color="000000")
        c.fill = fill(COLOR_LGREEN)
        c.alignment = center; c.border = border

    # --- Phase definitions ---
    phases = [
        ("Fresh Run\n(1M vectors, no delete)", "fresh", None, None),
    ]
    for rk in DELETE_RATES:
        exp = EXPECTED_COUNT_AFTER_DELETE[rk]
        phases.append((
            f"After Delete {rk}\n(verify count={exp:,})",
            f"post_delete_{rk}", f"delete_{rk}", None
        ))
        phases.append((
            f"After Reinsert {rk}\n(verify count={TOTAL_VECTORS:,})",
            f"post_insert_{rk}", None, f"insert_{rk}"
        ))

    # --- Data rows ---
    # Column order for metrics: 0.99, 0.99_prefilter, 0.80, 0.50, 0.01
    metric_keys = [0.99, "0.99_prefilter", 0.80, 0.50, 0.01]

    for i, (phase_name, bench_key, del_key, ins_key) in enumerate(phases):
        row = 3 + i
        ws.row_dimensions[row].height = 32
        row_fill = fill(COLOR_ROW_ODD) if i % 2 == 0 else fill("FFFFFF")

        c = ws.cell(row=row, column=1, value=phase_name)
        c.font = Font(bold=True); c.fill = row_fill
        c.alignment = left; c.border = border

        bench_results = all_data.get(bench_key) or {}
        col = 2
        for mk in metric_keys:
            m = bench_results.get(mk) or {}
            recall = m.get("recall")
            qps    = m.get("qps")
            p99    = m.get("p99_latency")

            vals = [
                round(recall * 100, 2) if recall is not None else "N/A",
                round(qps, 4)          if qps    is not None else "N/A",
                round(p99, 6)          if p99    is not None else "N/A",
            ]
            for v in vals:
                c = ws.cell(row=row, column=col, value=v)
                c.fill = row_fill; c.alignment = center; c.border = border
                col += 1

        del_dur = all_data.get(del_key) if del_key else None
        ins_dur = all_data.get(ins_key) if ins_key else None

        c = ws.cell(row=row, column=col,
                    value=round(del_dur, 2) if del_dur is not None else "")
        c.fill = row_fill; c.alignment = center; c.border = border

        c = ws.cell(row=row, column=col + 1,
                    value=round(ins_dur, 2) if ins_dur is not None else "")
        c.fill = row_fill; c.alignment = center; c.border = border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 30
    for col_letter in ["B","C","D","E","F","G","H","I","J","K","L","M","N","O","P"]:
        ws.column_dimensions[col_letter].width = 15
    ws.column_dimensions["Q"].width = 22
    ws.column_dimensions["R"].width = 22

    ws.freeze_panes = "B3"

    wb.save(output_path)
    print(f"\n[EXCEL] Saved: {output_path}")


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 65)
    print(" Stability Test — GT Filter Operator")
    print(f" Index : {INDEX_NAME}")
    print(f" Output: {OUTPUT_EXCEL}")
    print("=" * 65)

    client = EndeeClient(token=TOKEN)
    client.set_base_url(BASE_URL)
    index = client.get_index(INDEX_NAME)
    print(f"\n[INIT] Connected to index '{INDEX_NAME}'")
    print(f"[INIT] Current state: {index.describe()}")

    all_data: dict = {}

    # ---- Phase 1: Fresh run ----
    all_data["fresh"] = run_all_filter_rates("Fresh — 1M vectors")

    for rate_key in DELETE_RATES:
        # ---- Delete ----
        all_data[f"delete_{rate_key}"] = delete_vectors(index, rate_key)

        # Verify count after delete
        verify_count(index, expected=EXPECTED_COUNT_AFTER_DELETE[rate_key],
                     context=f"after delete {rate_key}")

        # ---- Bench post-delete ----
        all_data[f"post_delete_{rate_key}"] = run_all_filter_rates(
            f"Post-Delete {rate_key}"
        )

        # Wait before reinsert
        print(f"\n[WAIT] 10s pause before reinsert {rate_key} ...")
        time.sleep(10)

        # ---- Reinsert ----
        all_data[f"insert_{rate_key}"] = reinsert_vectors(index, rate_key)

        # Verify count after reinsert
        print("  [WAIT] 5s pause before verify count after reinsert ...")
        time.sleep(5)
        verify_count(index, expected=TOTAL_VECTORS,
                     context=f"after reinsert {rate_key}")

        # ---- Bench post-insert ----
        all_data[f"post_insert_{rate_key}"] = run_all_filter_rates(
            f"Post-Reinsert {rate_key}"
        )

        # Wait before next delete cycle (skip after last)
        if rate_key != DELETE_RATES[-1]:
            print(f"\n[WAIT] 10s pause before next delete cycle ...")
            time.sleep(10)

    # ---- Save results ----
    write_excel(all_data, OUTPUT_EXCEL)

    print("\n" + "=" * 65)
    print(" Stability test complete!")
    print(f" Results: {OUTPUT_EXCEL}")
    print("=" * 65)


if __name__ == "__main__":
    main()
